"""
Sales Dashboard Mockup — Professional Excel with:
- Correct US number formatting ($2,464,279.62 not $24,64,279.62)
- Frozen header rows
- Conditional colour on return rate and margin columns
- Clean chart placement (no overlaps)
- All 8 products visible
- Print-ready layout
"""

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── Colour palette ──────────────────────────────────────────────────────────
C_NAVY    = "0D1B2A"   # title background
C_TEAL    = "2F6F73"   # products header
C_BLUE    = "1B4F72"   # column headers
C_PURPLE  = "5368A6"   # campaign section
C_ORANGE  = "C06000"   # channel section
C_GREEN   = "1A7A4A"   # region section / good values
C_RED     = "C0392B"   # bad values
C_GOLD    = "F9A825"   # KPI accent
C_SLATE   = "455A64"   # column header bg
C_LGREY   = "EEF2F7"   # alternating row
C_WHITE   = "FFFFFF"
C_LBLUE   = "D6EAF8"   # light blue fill for totals

# ─── Number formats ───────────────────────────────────────────────────────────
F_USD  = '#,##0.00'     # 2,464,279.62
F_INT  = '#,##0'        # 1,250
F_PCT  = '0.00%'        # 20.32%
F_PCT1 = '0.0%'         # 20.3%

# ─── Border helpers ──────────────────────────────────────────────────────────
def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    t = Side(style="medium", color="888888")
    s = Side(style="thin",   color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=t)

# ─── Cell writer ─────────────────────────────────────────────────────────────
def w(ws, row, col, value, align="left", fmt=None,
      bold=False, bg=None, fg="111111", size=10, border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = thin()
    return c

def title(ws, row, text, span, bg=C_NAVY, fg=C_WHITE, size=15, h=34):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color=fg, size=size, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = h

def section(ws, row, col, text, span, bg, h=20):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=("  " + text))
    c.font      = Font(bold=True, color=C_WHITE, size=10, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin()
    ws.row_dimensions[row].height = h

def colhdr(ws, row, col, text, bg=C_SLATE, h=26):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color=C_WHITE, size=9, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    c.border    = thick_bottom()
    ws.row_dimensions[row].height = h

def kpi(ws, r_lbl, r_val, col, label, value, bg):
    # spans 2 columns
    for r, v, sz, bold in [(r_lbl, label, 8, True),
                            (r_val, value, 17, True)]:
        ws.merge_cells(start_row=r, start_column=col,
                       end_row=r,   end_column=col+1)
        c = ws.cell(row=r, column=col, value=v)
        c.font      = Font(bold=bold, color=C_WHITE, size=sz, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    # gold accent line
    ws.merge_cells(start_row=r_val+1, start_column=col,
                   end_row=r_val+1,   end_column=col+1)
    a = ws.cell(row=r_val+1, column=col, value="")
    a.fill = PatternFill("solid", fgColor=C_GOLD)

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Executive Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Dashboard"
ws1.sheet_view.showGridLines  = False
ws1.sheet_view.zoomScale      = 85
ws1.sheet_properties.tabColor = "2F6F73"

# Column widths  (A-K, F = narrow spacer)
widths = {"A":23,"B":16,"C":9,"D":15,"E":11,
          "F":2,
          "G":16,"H":16,"I":9,"J":16,"K":11}
for col, w_ in widths.items():
    ws1.column_dimensions[col].width = w_

# Row 1 — master title
title(ws1, 1, "SALES PERFORMANCE EXECUTIVE DASHBOARD", 11, h=36, size=16)
ws1.row_dimensions[2].height = 5   # tiny spacer

# Rows 3-5 — KPI cards  (2 cols each: A-B, C-D, E-F, G-H, I-J ... but F=spacer)
# Use cols: 1-2, 3-4, 5-6(gap), 7-8, 9-10
kpi_data = [
    (1,  "TOTAL REVENUE",    "$5,328,492",  C_TEAL),
    (3,  "TOTAL ORDERS",     "1,250",        C_PURPLE),
    (5,  "AVG ORDER VALUE",  "$4,262.79",   C_ORANGE),
    (7,  "GROSS MARGIN",     "20.32%",      C_GREEN),
    (9,  "RETURN RATE",      "5.20%",       C_RED),
]
ws1.row_dimensions[3].height = 16
ws1.row_dimensions[4].height = 36
ws1.row_dimensions[5].height = 4
for col, lbl, val, bg in kpi_data:
    kpi(ws1, 3, 4, col, lbl, val, bg)
ws1.row_dimensions[6].height = 8   # spacer

# ── LEFT PANEL: Cols A-E ────────────────────────────────────────────────────

# TOP PRODUCTS  rows 7-17
section(ws1, 7, 1, "TOP PRODUCTS BY REVENUE", 5, C_TEAL)
for i, h in enumerate(["Product Name","Revenue ($)","Orders","Avg Price ($)","Return %"]):
    colhdr(ws1, 8, i+1, h)

products = [
    ("Smart Watch",          1368209.84, 152, 6042.26, 0.0724),
    ("Mechanical Keyboard",   899189.76, 156, 3499.89, 0.0128),
    ("Webcam Pro",            798137.50, 169, 3170.67, 0.0473),
    ("Bluetooth Headphones",  655422.34, 151, 2654.19, 0.0530),
    ("Portable Speaker",      600264.00, 157, 2208.11, 0.0764),
    ("USB-C Hub",             506903.56, 161, 1911.33, 0.0683),
    ("Laptop Stand",          296637.20, 143, 1307.75, 0.0490),
    ("Wireless Mouse",        203728.20, 161,  798.97, 0.0373),
]
for r, (name, rev, ord_, avg, ret) in enumerate(products):
    row = 9 + r
    bg  = C_LGREY if r % 2 == 0 else C_WHITE
    ws1.row_dimensions[row].height = 16
    w(ws1, row, 1, name,  "left",   bg=bg)
    w(ws1, row, 2, rev,   "right",  F_USD, bg=bg)
    w(ws1, row, 3, ord_,  "center", F_INT, bg=bg)
    w(ws1, row, 4, avg,   "right",  F_USD, bg=bg)
    ret_bg = "FFE0E0" if ret > 0.065 else ("FFF9E0" if ret > 0.045 else "E8F8F5")
    w(ws1, row, 5, ret,   "center", F_PCT, bg=ret_bg)

ws1.row_dimensions[17].height = 8

# PERFORMANCE BY REGION  rows 18-23
section(ws1, 18, 1, "PERFORMANCE BY REGION", 5, C_GREEN)
for i, h in enumerate(["Region","Revenue ($)","Orders","Gross Profit ($)","Return %"]):
    colhdr(ws1, 19, i+1, h)

regions = [
    ("West",  1451298.87, 339, 273233.41, 0.0383),
    ("South", 1301485.12, 320, 250426.93, 0.0719),
    ("North", 1288064.81, 299, 315503.55, 0.0569),
    ("East",  1287643.60, 292, 243585.42, 0.0411),
]
for r, (reg, rev, ord_, gp, ret) in enumerate(regions):
    row = 20 + r
    bg  = C_LGREY if r % 2 == 0 else C_WHITE
    ws1.row_dimensions[row].height = 16
    w(ws1, row, 1, reg,  "left",   bg=bg)
    w(ws1, row, 2, rev,  "right",  F_USD, bg=bg)
    w(ws1, row, 3, ord_, "center", F_INT, bg=bg)
    w(ws1, row, 4, gp,   "right",  F_USD, bg=bg)
    ret_bg = "FFE0E0" if ret > 0.065 else ("FFF9E0" if ret > 0.045 else "E8F8F5")
    w(ws1, row, 5, ret,  "center", F_PCT, bg=ret_bg)

ws1.row_dimensions[24].height = 8

# ── RIGHT PANEL: Cols G-K ───────────────────────────────────────────────────

# REVENUE BY CHANNEL  rows 7-12
section(ws1, 7, 7, "REVENUE BY SALES CHANNEL", 5, C_ORANGE)
for i, h in enumerate(["Channel","Revenue ($)","Orders","Gross Profit ($)","Margin %"]):
    colhdr(ws1, 8, i+7, h)

channels = [
    ("Website",      2464279.62, 592, 488850.64, 0.1984),
    ("Mobile App",   2054498.79, 456, 441586.34, 0.2149),
    ("Retail Store",  809713.99, 202, 152312.33, 0.1881),
]
for r, (ch, rev, ord_, gp, margin) in enumerate(channels):
    row = 9 + r
    bg  = C_LGREY if r % 2 == 0 else C_WHITE
    ws1.row_dimensions[row].height = 16
    w(ws1, row, 7,  ch,     "left",   bg=bg)
    w(ws1, row, 8,  rev,    "right",  F_USD, bg=bg)
    w(ws1, row, 9,  ord_,   "center", F_INT, bg=bg)
    w(ws1, row, 10, gp,     "right",  F_USD, bg=bg)
    m_bg = "E8F8F5" if margin > 0.20 else C_LGREY
    w(ws1, row, 11, margin, "center", F_PCT, bg=m_bg)

ws1.row_dimensions[12].height = 8

# MARKETING CAMPAIGN  rows 13-21
section(ws1, 13, 7, "MARKETING CAMPAIGN SUMMARY", 5, C_PURPLE)
for i, h in enumerate(["Campaign","Revenue ($)","Orders","Gross Profit ($)","Margin %"]):
    colhdr(ws1, 14, i+7, h)

campaigns = [
    ("Search Ads",   1034918.19, 219, 276020.30, 0.2667),
    ("Organic",       946146.16, 217, 198461.91, 0.2098),
    ("Referral",      904176.23, 205, 185194.66, 0.2048),
    ("Email Offer",   844455.76, 222, 117325.28, 0.1389),
    ("Influencer",    809556.17, 191, 158623.78, 0.1959),
    ("Festive Sale",  789239.89, 196, 147123.38, 0.1864),
]
for r, (camp, rev, ord_, gp, margin) in enumerate(campaigns):
    row = 15 + r
    bg  = C_LGREY if r % 2 == 0 else C_WHITE
    ws1.row_dimensions[row].height = 16
    w(ws1, row, 7,  camp,   "left",   bg=bg)
    w(ws1, row, 8,  rev,    "right",  F_USD, bg=bg)
    w(ws1, row, 9,  ord_,   "center", F_INT, bg=bg)
    w(ws1, row, 10, gp,     "right",  F_USD, bg=bg)
    m_bg = "FFE0E0" if margin < 0.16 else ("E8F8F5" if margin > 0.24 else C_LGREY)
    w(ws1, row, 11, margin, "center", F_PCT, bg=m_bg)

# ── CHARTS on Sheet 1 (row 25+) ──────────────────────────────────────────────
# Pie: Revenue by Channel  → A25
pie1 = PieChart()
pie1.title  = "Revenue by Sales Channel"
pie1.style  = 26; pie1.width = 14; pie1.height = 12
pie1.add_data(Reference(ws1, min_col=8, min_row=8, max_row=11), titles_from_data=True)
pie1.set_categories(Reference(ws1, min_col=7, min_row=9, max_row=11))
for i, col in enumerate([C_TEAL, C_PURPLE, C_ORANGE]):
    pt = DataPoint(idx=i); pt.graphicalProperties.solidFill = col
    pie1.series[0].dPt.append(pt)
ws1.add_chart(pie1, "A25")

# Bar: Top Products  → G25
bar1 = BarChart()
bar1.type = "bar"; bar1.title = "Top 8 Products — Revenue ($)"
bar1.style = 26; bar1.width = 20; bar1.height = 14
bar1.add_data(Reference(ws1, min_col=2, min_row=8, max_row=16), titles_from_data=True)
bar1.set_categories(Reference(ws1, min_col=1, min_row=9, max_row=16))
bar1.series[0].graphicalProperties.solidFill      = C_TEAL
bar1.series[0].graphicalProperties.line.solidFill = C_TEAL
ws1.add_chart(bar1, "G25")

# Freeze pane below KPIs
ws1.freeze_panes = "A7"

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Monthly Trend
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Trend")
ws2.sheet_view.showGridLines  = False
ws2.sheet_view.zoomScale      = 85
ws2.sheet_properties.tabColor = "5368A6"

for col, w_ in {"A":12,"B":18,"C":10,"D":18,"E":18,"F":13}.items():
    ws2.column_dimensions[col].width = w_

title(ws2, 1, "MONTHLY SALES AND PROFIT TREND  |  Jan 2025 – Mar 2026", 6, h=30, size=13)
ws2.row_dimensions[2].height = 5

for i, h in enumerate(["Order Month","Total Revenue ($)","Orders",
                        "Total Cost ($)","Gross Profit ($)","Profit Margin %"]):
    colhdr(ws2, 3, i+1, h, bg="1B4F72")

monthly = [
    ("2025-01", 414508.30, 101, 336284.42,  78223.88, 0.1887),
    ("2025-02", 381797.56,  84, 282834.36,  98963.20, 0.2592),
    ("2025-03", 253183.26,  80, 204061.76,  49121.50, 0.1940),
    ("2025-04", 361993.81,  88, 299647.47,  62346.34, 0.1722),
    ("2025-05", 332373.01,  86, 270547.12,  61825.89, 0.1860),
    ("2025-06", 373196.49,  77, 305755.34,  67441.15, 0.1807),
    ("2025-07", 397458.33,  88, 318768.70,  78689.63, 0.1980),
    ("2025-08", 334842.66,  94, 275722.86,  59119.80, 0.1766),
    ("2025-09", 345674.85,  77, 279138.65,  66536.20, 0.1925),
    ("2025-10", 364085.70,  77, 292943.39,  71142.31, 0.1954),
    ("2025-11", 330265.79,  84, 264816.02,  65449.77, 0.1982),
    ("2025-12", 326024.85,  87, 264659.61,  61365.24, 0.1882),
    ("2026-01", 441892.25,  89, 307375.01, 134517.24, 0.3044),
    ("2026-02", 388410.22,  74, 317467.41,  70942.81, 0.1826),
    ("2026-03", 282785.32,  64, 225720.97,  57064.35, 0.2018),
]
for r, d in enumerate(monthly):
    row = 4 + r
    bg  = C_LGREY if r % 2 == 0 else C_WHITE
    ws2.row_dimensions[row].height = 16
    w(ws2, row, 1, d[0], "center", bg=bg)
    w(ws2, row, 2, d[1], "right",  F_USD, bg=bg)
    w(ws2, row, 3, d[2], "center", F_INT, bg=bg)
    w(ws2, row, 4, d[3], "right",  F_USD, bg=bg)
    w(ws2, row, 5, d[4], "right",  F_USD, bg=bg)
    m_bg = "E8F8F5" if d[5] > 0.25 else ("FFE0E0" if d[5] < 0.175 else bg)
    w(ws2, row, 6, d[5], "center", F_PCT, bg=m_bg)

# TOTAL row
tr = 19; ws2.row_dimensions[tr].height = 20
totals = [("TOTAL", None), (5328492.40, F_USD), (1250, F_INT),
          (4245778.09, F_USD), (1082548.82, F_USD), (0.2032, F_PCT)]
for i, (val, fmt_) in enumerate(totals):
    c = ws2.cell(row=tr, column=i+1, value=val)
    c.font      = Font(bold=True, color=C_WHITE, size=10, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="right" if i > 0 else "center",
                            vertical="center")
    if fmt_: c.number_format = fmt_
    c.border = thin()

ws2.row_dimensions[20].height = 8
ws2.freeze_panes = "A4"

# Line chart — Revenue vs Profit  A21
lc = LineChart()
lc.title = "Monthly Revenue vs Gross Profit"
lc.y_axis.title = "Amount ($)"; lc.x_axis.title = "Month"
lc.style = 26; lc.width = 28; lc.height = 14
lc.add_data(Reference(ws2, min_col=2, min_row=3, max_row=18), titles_from_data=True)
lc.add_data(Reference(ws2, min_col=5, min_row=3, max_row=18), titles_from_data=True)
lc.set_categories(Reference(ws2, min_col=1, min_row=4, max_row=18))
lc.series[0].graphicalProperties.line.solidFill = C_PURPLE
lc.series[0].graphicalProperties.line.width     = 28000
lc.series[1].graphicalProperties.line.solidFill = C_GREEN
lc.series[1].graphicalProperties.line.width     = 28000
ws2.add_chart(lc, "A21")

# Column chart — Order Count  A39
bc2 = BarChart()
bc2.type = "col"; bc2.title = "Monthly Order Count"
bc2.y_axis.title = "Orders"; bc2.x_axis.title = "Month"
bc2.style = 26; bc2.width = 28; bc2.height = 12
bc2.add_data(Reference(ws2, min_col=3, min_row=3, max_row=18), titles_from_data=True)
bc2.set_categories(Reference(ws2, min_col=1, min_row=4, max_row=18))
bc2.series[0].graphicalProperties.solidFill      = C_ORANGE
bc2.series[0].graphicalProperties.line.solidFill = C_ORANGE
ws2.add_chart(bc2, "A39")

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — Category & Segments
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Category & Segments")
ws3.sheet_view.showGridLines  = False
ws3.sheet_view.zoomScale      = 85
ws3.sheet_properties.tabColor = "C06000"

for col, w_ in {"A":14,"B":16,"C":9,"D":14,"E":16,"F":12,
                "G":2, "H":12,"I":16,"J":9,"K":16}.items():
    ws3.column_dimensions[col].width = w_

title(ws3, 1, "CATEGORY & SEGMENT BREAKDOWNS", 11, h=30, size=14)
ws3.row_dimensions[2].height = 5

# Category table A-F
section(ws3, 3, 1, "SALES BY PRODUCT CATEGORY", 6, C_TEAL)
for i, h in enumerate(["Category","Revenue ($)","Orders","Avg Price ($)","Gross Profit ($)","Margin %"]):
    colhdr(ws3, 4, i+1, h)

categories = [
    ("Accessories", 1609821.52, 478, 2055.11, 345746.17, 0.2148),
    ("Wearables",   1368209.84, 152, 6042.26, 208448.69, 0.1524),
    ("Audio",       1255686.34, 308, 2426.80, 264390.68, 0.2106),
    ("Office",      1094774.70, 312, 2316.83, 264163.77, 0.2413),
]
for r, (cat, rev, ord_, avg, gp, margin) in enumerate(categories):
    row = 5 + r
    bg  = C_LGREY if r % 2 == 0 else C_WHITE
    ws3.row_dimensions[row].height = 16
    w(ws3, row, 1, cat,    "left",   bg=bg)
    w(ws3, row, 2, rev,    "right",  F_USD, bg=bg)
    w(ws3, row, 3, ord_,   "center", F_INT, bg=bg)
    w(ws3, row, 4, avg,    "right",  F_USD, bg=bg)
    w(ws3, row, 5, gp,     "right",  F_USD, bg=bg)
    m_bg = "E8F8F5" if margin > 0.22 else ("FFE0E0" if margin < 0.16 else bg)
    w(ws3, row, 6, margin, "center", F_PCT, bg=m_bg)

# Segment table H-K
section(ws3, 3, 8, "SALES BY CUSTOMER SEGMENT", 4, C_PURPLE)
for i, h in enumerate(["Age Group","Revenue ($)","Orders","Rev / Customer ($)"]):
    colhdr(ws3, 4, i+8, h)

segments = [
    ("18–24", 884965.42,  210, 5746.53),
    ("25–34", 1032052.93, 246, 5639.63),
    ("35–44", 1206867.84, 268, 6004.32),
    ("45–54", 1059588.09, 267, 5271.58),
    ("55+",   1145018.12, 259, 6222.92),
]
for r, (age, rev, ord_, rpc) in enumerate(segments):
    row = 5 + r
    bg  = C_LGREY if r % 2 == 0 else C_WHITE
    ws3.row_dimensions[row].height = 16
    rpc_bg = "E8F8F5" if rpc > 6000 else bg
    w(ws3, row, 8,  age,  "center", bg=bg)
    w(ws3, row, 9,  rev,  "right",  F_USD, bg=bg)
    w(ws3, row, 10, ord_, "center", F_INT, bg=bg)
    w(ws3, row, 11, rpc,  "right",  F_USD, bg=rpc_bg)

ws3.row_dimensions[10].height = 8
ws3.freeze_panes = "A5"

# Bar: Category Revenue  A11
bc3 = BarChart()
bc3.type = "col"; bc3.title = "Revenue by Category ($)"
bc3.style = 26; bc3.width = 18; bc3.height = 12
bc3.add_data(Reference(ws3, min_col=2, min_row=4, max_row=8), titles_from_data=True)
bc3.set_categories(Reference(ws3, min_col=1, min_row=5, max_row=8))
bc3.series[0].graphicalProperties.solidFill      = C_TEAL
bc3.series[0].graphicalProperties.line.solidFill = C_TEAL
ws3.add_chart(bc3, "A11")

# Pie: Category Share  A28
pc2 = PieChart()
pc2.title = "Category Revenue Share"
pc2.style = 26; pc2.width = 14; pc2.height = 11
pc2.add_data(Reference(ws3, min_col=2, min_row=4, max_row=8), titles_from_data=True)
pc2.set_categories(Reference(ws3, min_col=1, min_row=5, max_row=8))
for i, col in enumerate([C_TEAL, C_PURPLE, C_ORANGE, C_GREEN]):
    pt = DataPoint(idx=i); pt.graphicalProperties.solidFill = col
    pc2.series[0].dPt.append(pt)
ws3.add_chart(pc2, "A28")

# Bar: Rev/Customer by Age  H11
bc4 = BarChart()
bc4.type = "col"; bc4.title = "Revenue per Customer by Age ($)"
bc4.style = 26; bc4.width = 18; bc4.height = 12
bc4.add_data(Reference(ws3, min_col=11, min_row=4, max_row=9), titles_from_data=True)
bc4.set_categories(Reference(ws3, min_col=8, min_row=5, max_row=9))
bc4.series[0].graphicalProperties.solidFill      = C_PURPLE
bc4.series[0].graphicalProperties.line.solidFill = C_PURPLE
ws3.add_chart(bc4, "H11")

# Bar: Total Revenue by Age  H28
bc5 = BarChart()
bc5.type = "col"; bc5.title = "Total Revenue by Age Group ($)"
bc5.style = 26; bc5.width = 18; bc5.height = 11
bc5.add_data(Reference(ws3, min_col=9, min_row=4, max_row=9), titles_from_data=True)
bc5.set_categories(Reference(ws3, min_col=8, min_row=5, max_row=9))
bc5.series[0].graphicalProperties.solidFill      = C_GREEN
bc5.series[0].graphicalProperties.line.solidFill = C_GREEN
ws3.add_chart(bc5, "H28")

# ─── Save ────────────────────────────────────────────────────────────────────
OUT = (r"c:\Users\lucky\Downloads\Lucky-DataAnalyst-Internship-Portfolio"
       r"\task-2-eda-business-intelligence\sales_dashboard_mockup.xlsx")
wb.save(OUT)
print(f"✅  Saved → {OUT}")
